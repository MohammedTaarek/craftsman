from flask import Flask, render_template, request, redirect, url_for
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook

app = Flask(__name__)
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 465
app.config["MAIL_USE_SSL"] = True
app.config["MAIL_USERNAME"] = "your_email"
app.config["MAIL_PASSWORD"] = "your_password"
app.config["SECRET_KEY"] = "your_secret_key"

mail = Mail(app)

users = {}

## Home Route
@app.route("/")
def home():
    """Renders the home page."""
    return render_template("index.html")

## User Registration Route
@app.route("/register", methods=["POST"])
def register():
    """Registers a new user."""
    first_name = request.form["first_name"]
    last_name = request.form["last_name"]
    username = request.form["username"]
    email = request.form["email"]
    password = generate_password_hash(request.form["password"])
    phone_number = request.form["phone_number"]
    user_type = request.form["user_type"]

    if email in users:
        return "Email already exists"

    users[email] = {
        "first_name": first_name,
        "last_name": last_name,
        "username": username,
        "password": password,
        "phone_number": phone_number,
        "user_type": user_type
    }

    # Save user data to file
    if user_type == "client":
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "First Name"
        ws["B1"] = "Last Name"
        ws["C1"] = "Username"
        ws["D1"] = "Email"
        ws["E1"] = "Password"
        ws["F1"] = "Phone Number"
        ws.append([first_name, last_name, username, email, password, phone_number])
        wb.save("clients.xlsx")
    elif user_type == "craftsman":
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "First Name"
        ws["B1"] = "Last Name"
        ws["C1"] = "Username"
        ws["D1"] = "Email"
        ws["E1"] = "Password"
        ws["F1"] = "Phone Number"
        ws.append([first_name, last_name, username, email, password, phone_number])
        wb.save("craftsmen.xlsx")

    # Send confirmation email to user after registration (your email ) is the email you used to set up the mail server in the app.config)
    msg = Message("Hello " + first_name, sender="your_email", recipients=[email])
    msg.body = "Thank you for registering!"
    mail.send(msg)

    return redirect(url_for("home"))

## Login Route
@app.route("/login", methods=["POST"])
def login():
    """Logs in an existing user."""
    email_or_phone = request.form["email_or_phone"]
    password = request.form["password"]

    for user in users.values():
        if user["email"] == email_or_phone or user["phone_number"] == email_or_phone:
            if check_password_hash(user["password"], password):
                return "Login successful"
            else:
                return "Incorrect password!"

    return "Email or phone number not found!"

if __name__ == "__main__":
    app.run(debug=True)



